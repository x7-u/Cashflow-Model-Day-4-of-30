"""Day 4. Multi-sheet Excel workbook for a cash flow analysis run.

Sheets:
1. Summary    Headline KPIs, embedded waterfall + balance line PNGs, run timestamp.
2. Forecast   Wide W1..W13 grid per line, totals column, RAG fills on closing.
3. Weekly     One row per week: receipts, payments, payroll, tax, net, opening, closing, RAG.
4. Scenario   Only present if a scenario has been run. Base vs scenario per week with delta.
5. Commentary Only present after a scenario. AI summary, interpretation, actions, cost stats.
6. Inputs     Echo of parsed input lines for reproducibility.

Currency format: '"£"#,##0.00;[Red]-"£"#,##0.00'.
RAG fills: green / amber / red from shared.chart_styles.RAG_HEX, plus a
'favourable' blue for forecast totals if any line ended favourable, and an
'na' grey for empty rows.
"""
from __future__ import annotations

import datetime as dt
import sys
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from chart import render_balance_line_png, render_waterfall_png
from pipeline import AnalysisResult, ScenarioResult

from shared.chart_styles import RAG_HEX

CURRENCY_FMT = '"£"#,##0.00;[Red]-"£"#,##0.00'

GREEN_HEX = RAG_HEX.get("green", "9DD9A2").lstrip("#")
AMBER_HEX = RAG_HEX.get("amber", "F2C76A").lstrip("#")
RED_HEX   = RAG_HEX.get("red",   "F08A8A").lstrip("#")
BLUE_HEX  = "B7CEF0"
NA_HEX    = "B6BCC6"
PANEL_HEX = "EFF4F7"
INK_HEX   = "0A1019"
TEAL_HEX  = "2D9CA5"

THIN = Side(style="thin", color="C9CFD9")
THIN_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BOLD = Font(bold=True, color=INK_HEX)
TEAL_BOLD = Font(bold=True, color=TEAL_HEX)


def _slug(s: str) -> str:
    out = "".join(ch.lower() if ch.isalnum() else "_" for ch in s)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_") or "company"


def write_workbook(
    result: AnalysisResult,
    out_dir: Path,
    *,
    scenario: ScenarioResult | None = None,
) -> Path:
    """Write the workbook to ``out_dir`` and return the path."""
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, result, scenario=scenario)
    _write_forecast(wb, result)
    _write_weekly(wb, result)
    if scenario is not None:
        _write_scenario(wb, scenario)
        _write_commentary(wb, scenario)
    _write_inputs(wb, result)

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M")
    name = f"cashflow_{_slug(result.metadata.company)}_{_slug(result.metadata.period_label)}_{stamp}.xlsx"
    path = out_dir / name
    wb.save(path)
    return path


# ---- Summary ---------------------------------------------------------

def _write_summary(wb, result: AnalysisResult, *, scenario: ScenarioResult | None) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    ws["A1"] = "Cash Flow Forecast"
    ws["A1"].font = Font(bold=True, size=18, color=INK_HEX)
    ws["A2"] = f"{result.metadata.company} . {result.metadata.period_label}"
    ws["A2"].font = Font(italic=True, color=INK_HEX)
    ws["A3"] = f"Generated {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(color=INK_HEX, size=10)

    headline = result.headline
    rows: list[tuple[str, float | int | str | None, str | None]] = [
        ("Opening balance",       headline.opening_balance,    "currency"),
        ("Closing balance",       headline.closing_balance,    "currency"),
        ("Min balance",           headline.min_balance,        "currency"),
        ("Min balance week",      f"W{headline.min_balance_week}", None),
        ("Max balance",           headline.max_balance,        "currency"),
        ("Max balance week",      f"W{headline.max_balance_week}", None),
        ("Weeks below zero",      headline.weeks_below_zero,   None),
        ("Weeks below buffer",    headline.weeks_below_buffer, None),
        ("Runway weeks",          ("beyond horizon" if headline.runway_weeks is None
                                   else f"W{headline.runway_weeks}"), None),
        ("Total inflow",          headline.total_inflow,       "currency"),
        ("Total outflow",         headline.total_outflow,      "currency"),
        ("Cumulative net",        headline.cumulative_net,     "currency"),
        ("Buffer used",           headline.buffer_used,        "currency"),
        ("Concentration flags",   headline.concentration_flag_count, None),
    ]
    start = 5
    for i, (label, value, fmt) in enumerate(rows):
        ws.cell(row=start + i, column=1, value=label).font = BOLD
        c = ws.cell(row=start + i, column=2, value=value)
        if fmt == "currency":
            c.number_format = CURRENCY_FMT

    # Embed the chart PNGs.
    waterfall_bytes = render_waterfall_png(
        result.weeks,
        title=f"Waterfall {result.metadata.company} {result.metadata.period_label}",
    )
    line_bytes = render_balance_line_png(
        result.weeks,
        title="Closing balance over 13 weeks",
        buffer=headline.buffer_used,
        runway_week=headline.runway_weeks,
    )
    _embed_image(ws, waterfall_bytes, anchor="D5")
    _embed_image(ws, line_bytes, anchor="D29")

    if scenario is not None:
        s_row = start + len(rows) + 2
        ws.cell(row=s_row, column=1, value="Scenario applied").font = TEAL_BOLD
        ws.cell(row=s_row + 1, column=1, value="Interpretation").font = BOLD
        ws.cell(row=s_row + 1, column=2, value=scenario.commentary.interpretation)


# ---- Forecast (wide W1..W13) ----------------------------------------

def _write_forecast(wb, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Forecast")
    headers = ["name", "type", "category", *[f"W{i}" for i in range(1, 14)], "total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = BOLD
        cell.fill = PatternFill("solid", fgColor=PANEL_HEX)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for r, line in enumerate(result.lines, start=2):
        ws.cell(row=r, column=1, value=line.name)
        ws.cell(row=r, column=2, value=line.type)
        ws.cell(row=r, column=3, value=line.category)
        for w in range(1, 14):
            c = ws.cell(row=r, column=3 + w, value=line.amounts[w - 1])
            c.number_format = CURRENCY_FMT
        total_cell = ws.cell(row=r, column=17, value=line.total())
        total_cell.number_format = CURRENCY_FMT
        total_cell.font = BOLD

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 18
    for col in range(4, 17):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["Q"].width = 14

    # Autofilter on the header row for native drill-through.
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "D2"


# ---- Weekly ---------------------------------------------------------

def _write_weekly(wb, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Weekly")
    headers = [
        "week", "iso_week_label", "receipts", "payments", "payroll", "tax",
        "inflow", "outflow", "net", "opening", "closing", "rag", "concentration_pct",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor=PANEL_HEX)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    fills = {
        "green": PatternFill("solid", fgColor=GREEN_HEX),
        "amber": PatternFill("solid", fgColor=AMBER_HEX),
        "red":   PatternFill("solid", fgColor=RED_HEX),
    }
    currency_cols = {3, 4, 5, 6, 7, 8, 9, 10, 11}

    for r, w in enumerate(result.weeks, start=2):
        ws.cell(row=r, column=1, value=w.week)
        ws.cell(row=r, column=2, value=w.iso_week_label)
        ws.cell(row=r, column=3, value=w.receipts).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=w.payments).number_format = CURRENCY_FMT
        ws.cell(row=r, column=5, value=w.payroll).number_format = CURRENCY_FMT
        ws.cell(row=r, column=6, value=w.tax).number_format = CURRENCY_FMT
        ws.cell(row=r, column=7, value=w.inflow).number_format = CURRENCY_FMT
        ws.cell(row=r, column=8, value=w.outflow).number_format = CURRENCY_FMT
        ws.cell(row=r, column=9, value=w.net).number_format = CURRENCY_FMT
        ws.cell(row=r, column=10, value=w.opening).number_format = CURRENCY_FMT
        closing_cell = ws.cell(row=r, column=11, value=w.closing)
        closing_cell.number_format = CURRENCY_FMT
        closing_cell.fill = fills.get(w.rag, PatternFill())
        closing_cell.font = Font(bold=True, color="FFFFFF" if w.rag != "green" else INK_HEX)
        ws.cell(row=r, column=12, value=w.rag)
        if w.concentration_pct is not None:
            ws.cell(row=r, column=13, value=w.concentration_pct).number_format = "0.0%"
        for col in currency_cols:
            ws.cell(row=r, column=col).border = THIN_BORDER

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 14
    for col in range(3, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["L"].width = 9
    ws.column_dimensions["M"].width = 16
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"


# ---- Scenario sheet --------------------------------------------------

def _write_scenario(wb, scenario: ScenarioResult) -> None:
    ws = wb.create_sheet("Scenario")
    headers = [
        "week", "iso_week_label",
        "base_closing", "scenario_closing", "delta",
        "base_rag", "scenario_rag",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor=PANEL_HEX)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    fills = {
        "green": PatternFill("solid", fgColor=GREEN_HEX),
        "amber": PatternFill("solid", fgColor=AMBER_HEX),
        "red":   PatternFill("solid", fgColor=RED_HEX),
    }

    base_weeks = scenario.base.weeks
    sc_weeks = scenario.scenario.weeks
    for r, (b, s) in enumerate(zip(base_weeks, sc_weeks), start=2):
        ws.cell(row=r, column=1, value=b.week)
        ws.cell(row=r, column=2, value=b.iso_week_label)
        ws.cell(row=r, column=3, value=b.closing).number_format = CURRENCY_FMT
        sc_cell = ws.cell(row=r, column=4, value=s.closing)
        sc_cell.number_format = CURRENCY_FMT
        sc_cell.fill = fills.get(s.rag, PatternFill())
        ws.cell(row=r, column=5, value=s.closing - b.closing).number_format = CURRENCY_FMT
        ws.cell(row=r, column=6, value=b.rag)
        ws.cell(row=r, column=7, value=s.rag)

    for col in range(3, 6):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 14
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"


# ---- Commentary sheet -----------------------------------------------

def _write_commentary(wb, scenario: ScenarioResult) -> None:
    ws = wb.create_sheet("Commentary")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    c = scenario.commentary

    rows: list[tuple[str, str]] = [
        ("User prompt", scenario.user_prompt),
        ("Interpretation", c.interpretation),
        ("Headline", c.headline),
        ("Summary", c.summary),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = BOLD
        ws.cell(row=i, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = max(18, min(120, 14 * (1 + len(str(value)) // 80)))

    actions_start = len(rows) + 2
    ws.cell(row=actions_start, column=1, value="Actions").font = BOLD
    for j, action in enumerate(c.actions, start=actions_start):
        ws.cell(row=j, column=2, value=f". {action}")

    cost_row = actions_start + len(c.actions) + 2
    ws.cell(row=cost_row,     column=1, value="Cost USD").font = BOLD
    ws.cell(row=cost_row,     column=2, value=round(c.cost_usd, 6))
    ws.cell(row=cost_row + 1, column=1, value="Input tokens").font = BOLD
    ws.cell(row=cost_row + 1, column=2, value=c.input_tokens)
    ws.cell(row=cost_row + 2, column=1, value="Output tokens").font = BOLD
    ws.cell(row=cost_row + 2, column=2, value=c.output_tokens)
    ws.cell(row=cost_row + 3, column=1, value="Cache hits").font = BOLD
    ws.cell(row=cost_row + 3, column=2, value=c.cache_hit_tokens)
    ws.cell(row=cost_row + 4, column=1, value="Model").font = BOLD
    ws.cell(row=cost_row + 4, column=2, value=c.model)


# ---- Inputs echo ----------------------------------------------------

def _write_inputs(wb, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "Metadata"
    ws["A1"].font = TEAL_BOLD
    md = result.metadata
    pairs = [
        ("company", md.company),
        ("currency", md.currency),
        ("opening_balance", md.opening_balance),
        ("start_date", md.start_date.isoformat()),
        ("period_label", md.period_label),
        ("buffer_amount", md.buffer_amount if md.buffer_amount is not None else ""),
    ]
    for i, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)

    headers_row = len(pairs) + 4
    ws.cell(row=headers_row, column=1, value="Lines").font = TEAL_BOLD
    headers = ["name", "type", "category", *[f"W{i}" for i in range(1, 14)], "source"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=headers_row + 1, column=col, value=h)
        c.font = BOLD
        c.fill = PatternFill("solid", fgColor=PANEL_HEX)
    for r, line in enumerate(result.lines, start=headers_row + 2):
        ws.cell(row=r, column=1, value=line.name)
        ws.cell(row=r, column=2, value=line.type)
        ws.cell(row=r, column=3, value=line.category)
        for w in range(1, 14):
            c = ws.cell(row=r, column=3 + w, value=line.amounts[w - 1])
            c.number_format = CURRENCY_FMT
        ws.cell(row=r, column=17, value=line.source)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    for col in range(4, 17):
        ws.column_dimensions[get_column_letter(col)].width = 12

    if result.warnings:
        warn_row = headers_row + 2 + len(result.lines) + 2
        ws.cell(row=warn_row, column=1, value="Warnings").font = TEAL_BOLD
        for i, w in enumerate(result.warnings, start=warn_row + 1):
            ws.cell(row=i, column=1, value=w).alignment = Alignment(wrap_text=True)


# ---- Image embed helper ---------------------------------------------

def _embed_image(ws, png_bytes: bytes, *, anchor: str) -> None:
    """Embed a PNG into a sheet using a BytesIO buffer (file stays unopened)."""
    buf = BytesIO(png_bytes)
    img = XLImage(buf)
    img.width = 720
    img.height = 280
    ws.add_image(img, anchor)
