"""Day 4. Schema parsing tests.

Covers the canonical wide layout, the long-format adapter, the actuals
overlay, GBP enforcement, missing keys and ambiguity rejection.
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from cashflow_schema import (
    WEEK_COUNT,
    CashflowLine,
    parse_inputs,
    week_iso_label,
)

DAY_ROOT = Path(__file__).resolve().parent.parent
SAMPLES = DAY_ROOT / "sample_data"


def _make_workbook(metadata_pairs, *, forecast_rows=None, data_rows=None,
                   actuals_rows=None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    md = wb.create_sheet("metadata")
    md.append(["key", "value"])
    for k, v in metadata_pairs:
        md.append([k, v])

    if forecast_rows is not None:
        fc = wb.create_sheet("forecast")
        fc.append(["name", "type", "category", *[f"W{i}" for i in range(1, 14)]])
        for row in forecast_rows:
            fc.append(row)

    if data_rows is not None:
        dt = wb.create_sheet("data")
        dt.append(["week", "name", "type", "category", "amount", "source"])
        for row in data_rows:
            dt.append(row)

    if actuals_rows is not None:
        ac = wb.create_sheet("actuals")
        ac.append(["name", "type", "category", *[f"W{i}" for i in range(1, 14)]])
        for row in actuals_rows:
            ac.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_MD_OK = [
    ("company", "Test Co"),
    ("currency", "GBP"),
    ("opening_balance", 10000),
    ("start_date", "2026-04-06"),
    ("period_label", "Q2 26"),
]


def test_required_metadata_keys_enforced() -> None:
    data = _make_workbook(
        [("company", "X"), ("currency", "GBP"), ("opening_balance", 0)],
        forecast_rows=[["Sales", "receipt", "AR", 100, *([0] * 12)]],
    )
    with pytest.raises(ValueError, match="missing keys"):
        parse_inputs(file_bytes=data, source_filename="x.xlsx")


def test_gbp_only_in_mvp() -> None:
    md = list(_MD_OK)
    md[1] = ("currency", "USD")
    data = _make_workbook(md, forecast_rows=[["Sales", "receipt", "AR", 100, *([0] * 12)]])
    with pytest.raises(ValueError, match="GBP only"):
        parse_inputs(file_bytes=data, source_filename="x.xlsx")


def test_wide_layout_happy_path() -> None:
    data = _make_workbook(
        _MD_OK,
        forecast_rows=[
            ["Sales", "receipt", "AR", 1000, 1100, 1200, *([0] * 10)],
            ["Rent",  "payment", "rent", 500, 500, 500, *([0] * 10)],
        ],
    )
    parsed = parse_inputs(file_bytes=data, source_filename="x.xlsx")
    assert parsed.metadata.company == "Test Co"
    assert parsed.metadata.opening_balance == 10000.0
    assert len(parsed.lines) == 2
    assert parsed.lines[0].type == "receipt"
    assert parsed.lines[0].amounts[:3] == [1000.0, 1100.0, 1200.0]


def test_long_layout_adapter() -> None:
    data = _make_workbook(
        _MD_OK,
        data_rows=[
            [1, "Sales", "receipt", "AR", 1000, "forecast"],
            [2, "Sales", "receipt", "AR", 1100, "forecast"],
            [1, "Rent",  "payment", "rent",  500, "forecast"],
        ],
    )
    parsed = parse_inputs(file_bytes=data, source_filename="x.xlsx")
    assert len(parsed.lines) == 2
    sales = next(l for l in parsed.lines if l.name == "Sales")
    assert sales.amounts[0] == 1000.0
    assert sales.amounts[1] == 1100.0


def test_reject_when_both_layouts_present() -> None:
    data = _make_workbook(
        _MD_OK,
        forecast_rows=[["Sales", "receipt", "AR", 100, *([0] * 12)]],
        data_rows=[[1, "Sales", "receipt", "AR", 100, "forecast"]],
    )
    with pytest.raises(ValueError, match="both"):
        parse_inputs(file_bytes=data, source_filename="x.xlsx")


def test_actuals_overlay_overrides_forecast() -> None:
    data = _make_workbook(
        _MD_OK,
        forecast_rows=[["Sales", "receipt", "AR", 1000, 1000, 1000, *([0] * 10)]],
        actuals_rows=[["Sales", "receipt", "AR", 1500, 0, 0, *([0] * 10)]],
    )
    parsed = parse_inputs(file_bytes=data, source_filename="x.xlsx")
    sales = next(l for l in parsed.lines if l.name == "Sales")
    # W1 should be overridden, W2 and W3 keep the forecast value.
    assert sales.amounts[0] == 1500.0
    assert sales.amounts[1] == 1000.0
    assert sales.amounts[2] == 1000.0
    assert sales.source == "hybrid"


def test_unknown_line_type_warns_and_skips() -> None:
    data = _make_workbook(
        _MD_OK,
        forecast_rows=[["Mystery", "wibble", "?", 100, *([0] * 12)],
                       ["Sales",   "receipt", "AR", 100, *([0] * 12)]],
    )
    parsed = parse_inputs(file_bytes=data, source_filename="x.xlsx")
    assert len(parsed.lines) == 1
    assert any("unknown type" in w for w in parsed.warnings)


def test_empty_workbook_raises() -> None:
    data = _make_workbook(_MD_OK)  # no forecast, no data
    with pytest.raises(ValueError, match="neither"):
        parse_inputs(file_bytes=data, source_filename="x.xlsx")


def test_iso_week_label_helper() -> None:
    import datetime as dt
    label = week_iso_label(dt.date(2026, 4, 6), 1)  # Mon 2026-04-06
    assert label.startswith("2026-W")


def test_real_sample_round_trips() -> None:
    """One end-to-end check against the bundled services sample."""
    services = SAMPLES / "sample_services.xlsx"
    if not services.is_file():
        pytest.skip("services sample not built yet; run sample_data/_build.py first")
    parsed = parse_inputs(path=services, source_filename=services.name)
    assert parsed.metadata.company == "Acme Consulting Ltd"
    assert parsed.metadata.currency == "GBP"
    assert all(len(l.amounts) == WEEK_COUNT for l in parsed.lines)
    assert any(isinstance(l, CashflowLine) and l.type == "payroll" for l in parsed.lines)
