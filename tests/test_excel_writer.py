"""Day 4. Excel writer tests.

Smoke-level: write the workbook for the services sample, verify the
expected sheets, that the embedded chart PNGs are present, and that the
GBP currency format is applied to the right cells.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from excel_writer import CURRENCY_FMT, write_workbook
from pipeline import analyse

DAY_ROOT = Path(__file__).resolve().parent.parent
SAMPLES = DAY_ROOT / "sample_data"


@pytest.fixture(scope="module")
def services_result():
    services = SAMPLES / "sample_services.xlsx"
    if not services.is_file():
        pytest.skip("services sample not built; run sample_data/_build.py first")
    return analyse(path=services, source_filename=services.name, skip_ai=True)


def test_workbook_has_expected_sheets(services_result, tmp_path) -> None:
    out = write_workbook(services_result, tmp_path)
    wb = openpyxl.load_workbook(out)
    expected = ["Summary", "Forecast", "Weekly", "Inputs"]
    for s in expected:
        assert s in wb.sheetnames, f"missing sheet: {s}"


def test_workbook_embeds_chart_images(services_result, tmp_path) -> None:
    out = write_workbook(services_result, tmp_path)
    wb = openpyxl.load_workbook(out)
    summary = wb["Summary"]
    # openpyxl exposes embedded images on the sheet's _images list.
    images = getattr(summary, "_images", [])
    assert len(images) >= 2, f"expected 2 embedded PNGs on Summary, got {len(images)}"


def test_weekly_sheet_uses_currency_format(services_result, tmp_path) -> None:
    out = write_workbook(services_result, tmp_path)
    wb = openpyxl.load_workbook(out)
    weekly = wb["Weekly"]
    # Closing column should carry the GBP format on every data row.
    for r in range(2, 2 + len(services_result.weeks)):
        cell = weekly.cell(row=r, column=11)
        assert cell.number_format == CURRENCY_FMT, f"row {r} closing has wrong format"


def test_forecast_sheet_has_autofilter(services_result, tmp_path) -> None:
    out = write_workbook(services_result, tmp_path)
    wb = openpyxl.load_workbook(out)
    forecast = wb["Forecast"]
    assert forecast.auto_filter.ref is not None
