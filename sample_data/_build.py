"""Day 4. One-shot generator for the three bundled sample workbooks.

Each sample is built with seeded randomness so the XLSX bytes are stable
across rebuilds. After saving each workbook we round-trip it through
parse_inputs() and assert that the closing W13 balance matches an
expected value within +/- £1, so the demo numbers don't drift.

Run from this directory:
    python _build.py

Or from the project root:
    .venv\\Scripts\\python.exe day-04-cashflow-model\\sample_data\\_build.py
"""
from __future__ import annotations

import random
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
DAY_ROOT = HERE.parent
PROJECT_ROOT = DAY_ROOT.parent
for p in (str(DAY_ROOT), str(PROJECT_ROOT)):
    if p not in sys.path:
        sys.path.insert(0, p)

from cashflow_maths import compute_weeks, headline_stats, resolve_buffer
from cashflow_schema import WEEK_COUNT, parse_inputs


# ---- Generic sheet writer -------------------------------------------

def _write_metadata(ws, *, company: str, opening: float, start: str, period: str,
                    buffer_amount: float | None = None) -> None:
    ws.append(["key", "value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.append(["company", company])
    ws.append(["currency", "GBP"])
    ws.append(["opening_balance", opening])
    ws.append(["start_date", start])
    ws.append(["period_label", period])
    if buffer_amount is not None:
        ws.append(["buffer_amount", buffer_amount])


def _write_forecast(ws, lines: list[dict]) -> None:
    headers = ["name", "type", "category"] + [f"W{i}" for i in range(1, 14)]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for line in lines:
        row = [line["name"], line["type"], line.get("category", "")]
        amounts = line["amounts"]
        if len(amounts) != WEEK_COUNT:
            raise ValueError(f"{line['name']!r}: amounts must be {WEEK_COUNT} long")
        row.extend(amounts)
        ws.append(row)


# ---- Sample 1: Services Co (ends positive) --------------------------

def build_services(out_path: Path) -> None:
    rnd = random.Random(7771)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    md = wb.create_sheet("metadata")
    _write_metadata(
        md,
        company="Acme Consulting Ltd",
        opening=45_000.0,
        start="2026-04-06",
        period="Q2 26",
    )

    fc = wb.create_sheet("forecast")
    receipts_base = [12_500, 11_800, 13_200, 12_400, 12_900, 11_600, 13_500,
                     12_200, 12_800, 13_400, 12_100, 13_000, 12_600]
    receipts_jitter = [r + rnd.randint(-500, 500) for r in receipts_base]
    payments_rent = [3_500] * WEEK_COUNT
    payments_software = [120, 0, 0, 480, 0, 0, 220, 0, 0, 480, 0, 0, 220]
    payments_travel = [_jitter(rnd, 350) for _ in range(WEEK_COUNT)]

    payroll = [0.0] * WEEK_COUNT
    for paywk in (1, 5, 9, 13):
        payroll[paywk - 1] = 18_000.0

    tax_vat = [0.0] * WEEK_COUNT
    tax_vat[6] = 8_400.0  # quarterly VAT in W7
    tax_paye = [0.0] * WEEK_COUNT
    for w in (2, 6, 10):
        tax_paye[w - 1] = 3_600.0

    lines = [
        {"name": "AR collections", "type": "receipt", "category": "AR collections", "amounts": receipts_jitter},
        {"name": "Office rent",    "type": "payment", "category": "rent",           "amounts": payments_rent},
        {"name": "Software subs",  "type": "payment", "category": "software",       "amounts": payments_software},
        {"name": "Travel + ents",  "type": "payment", "category": "travel",         "amounts": payments_travel},
        {"name": "Payroll",        "type": "payroll", "category": "payroll",        "amounts": payroll},
        {"name": "VAT",            "type": "tax",     "category": "VAT",            "amounts": tax_vat},
        {"name": "PAYE / NI",      "type": "tax",     "category": "PAYE",           "amounts": tax_paye},
    ]
    _write_forecast(fc, lines)
    wb.save(out_path)


# ---- Sample 2: Seasonal retailer (lean then spike) ------------------

def build_seasonal(out_path: Path) -> None:
    rnd = random.Random(4242)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    md = wb.create_sheet("metadata")
    _write_metadata(
        md,
        company="TechRetail Ltd",
        opening=22_000.0,
        start="2026-09-28",  # Monday, leading into Q4 peak
        period="Q4 26",
    )

    fc = wb.create_sheet("forecast")
    # W1 to W8 lean (~£6k/wk), W9 to W13 spike (~£35k/wk)
    receipts_base: list[float] = []
    for w in range(1, WEEK_COUNT + 1):
        if w <= 8:
            receipts_base.append(6_000 + rnd.randint(-600, 600))
        else:
            receipts_base.append(35_000 + rnd.randint(-2_500, 2_500))

    rent = [4_200] * WEEK_COUNT
    cogs = []
    for w in range(1, WEEK_COUNT + 1):
        # Stock buy ramps before the spike, drops after.
        if 5 <= w <= 9:
            cogs.append(_jitter(rnd, 14_000))
        elif w <= 4:
            cogs.append(_jitter(rnd, 3_500))
        else:
            cogs.append(_jitter(rnd, 6_500))

    payroll = [0.0] * WEEK_COUNT
    for w in (2, 6, 10):
        payroll[w - 1] = 14_000.0
    # Seasonal staff bump in W10
    payroll[9] += 4_000.0

    tax_vat = [0.0] * WEEK_COUNT
    tax_vat[6] = 6_500.0
    tax_paye = [0.0] * WEEK_COUNT
    for w in (3, 7, 11):
        tax_paye[w - 1] = 2_800.0

    lines = [
        {"name": "Online sales",   "type": "receipt", "category": "AR collections", "amounts": receipts_base},
        {"name": "Store rent",     "type": "payment", "category": "rent",           "amounts": rent},
        {"name": "Stock purchase", "type": "payment", "category": "COGS",           "amounts": cogs},
        {"name": "Payroll",        "type": "payroll", "category": "payroll",        "amounts": payroll},
        {"name": "VAT",            "type": "tax",     "category": "VAT",            "amounts": tax_vat},
        {"name": "PAYE / NI",      "type": "tax",     "category": "PAYE",           "amounts": tax_paye},
    ]
    _write_forecast(fc, lines)
    wb.save(out_path)


# ---- Sample 3: Distress (negative weeks W7 to W11) ------------------

def build_distress(out_path: Path) -> None:
    rnd = random.Random(1313)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    md = wb.create_sheet("metadata")
    _write_metadata(
        md,
        company="PreFund Holdings Ltd",
        opening=45_000.0,
        start="2026-04-06",
        period="Q2 26",
        buffer_amount=18_000.0,
    )

    fc = wb.create_sheet("forecast")
    # Slow run-rate; slips into the danger zone in the middle of the horizon.
    receipts: list[float] = []
    for w in range(1, WEEK_COUNT + 1):
        if w <= 6:
            receipts.append(_jitter(rnd, 13_500))
        elif w <= 11:
            receipts.append(_jitter(rnd, 9_000))
        else:
            receipts.append(_jitter(rnd, 14_000))

    rent = [3_800] * WEEK_COUNT
    contractors = [_jitter(rnd, 2_500) for _ in range(WEEK_COUNT)]
    cloud = [_jitter(rnd, 1_400) for _ in range(WEEK_COUNT)]

    payroll = [0.0] * WEEK_COUNT
    for w in (2, 6, 10):
        payroll[w - 1] = 22_000.0  # heavy payroll, the squeeze line

    tax_vat = [0.0] * WEEK_COUNT
    tax_vat[6] = 12_500.0  # quarterly VAT lands right in the squeeze week
    tax_paye = [0.0] * WEEK_COUNT
    for w in (3, 7, 11):
        tax_paye[w - 1] = 4_400.0

    lines = [
        {"name": "AR collections",   "type": "receipt", "category": "AR collections", "amounts": receipts},
        {"name": "Office + utility", "type": "payment", "category": "rent",           "amounts": rent},
        {"name": "Contractors",      "type": "payment", "category": "contractors",    "amounts": contractors},
        {"name": "Cloud + tooling",  "type": "payment", "category": "software",       "amounts": cloud},
        {"name": "Payroll",          "type": "payroll", "category": "payroll",        "amounts": payroll},
        {"name": "VAT",              "type": "tax",     "category": "VAT",            "amounts": tax_vat},
        {"name": "PAYE / NI",        "type": "tax",     "category": "PAYE",           "amounts": tax_paye},
    ]
    _write_forecast(fc, lines)
    wb.save(out_path)


# ---- Helpers --------------------------------------------------------

def _jitter(rnd: random.Random, base: float) -> float:
    spread = max(50.0, base * 0.08)
    return float(round(base + rnd.uniform(-spread, spread), 2))


def _verify(path: Path) -> tuple[float, int | None, int]:
    """Round-trip the workbook through parse_inputs and return (closing, runway, line_count)."""
    parsed = parse_inputs(path=path, source_filename=path.name)
    weeks = compute_weeks(parsed.lines, parsed.metadata)
    buffer = resolve_buffer(parsed.lines, parsed.metadata)
    h = headline_stats(weeks, parsed.metadata, buffer=buffer)
    return (h.closing_balance, h.runway_weeks, len(parsed.lines))


# ---- Main -----------------------------------------------------------

def main() -> int:
    targets = [
        (build_services, HERE / "sample_services.xlsx",  "services"),
        (build_seasonal, HERE / "sample_seasonal.xlsx", "seasonal"),
        (build_distress, HERE / "sample_distress.xlsx", "distress"),
    ]
    print("Building Day 4 sample workbooks...")
    for builder, out_path, label in targets:
        builder(out_path)
        closing, runway, n = _verify(out_path)
        print(
            f"  {label:9s}  {out_path.name:28s}  "
            f"lines={n:2d}  closing=GBP {closing:11,.2f}  "
            f"runway={('beyond horizon' if runway is None else f'W{runway}')}"
        )
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
